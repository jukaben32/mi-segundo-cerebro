"""
convert_csv_to_xlsx.py
Converts any CSV file to a professionally formatted Excel spreadsheet.

Usage:
    python Scripts/convert_csv_to_xlsx.py "path/to/file.csv"
    python Scripts/convert_csv_to_xlsx.py "path/to/file.csv" "path/to/output.xlsx"

Requires: pip install pandas openpyxl
"""

import sys
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


# ── Palette ──────────────────────────────────────────────────────────────────
DARK_NAVY  = "1B2A47"
ACCENT     = "2E5F8A"
LIGHT_GRAY = "F4F6F8"
WHITE      = "FFFFFF"
MID_GRAY   = "C5CDD6"
TEXT_DARK  = "1C2B3A"

THIN_BORDER = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)

MIN_COL_WIDTH = 14
MAX_COL_WIDTH = 52
ROW_HEIGHT    = 18
HEADER_HEIGHT = 28
TITLE_HEIGHT  = 24
FONT_NAME     = "Calibri"


def strip_emojis(text: str) -> str:
    emoji_pattern = re.compile(
        "["
        u"\U0001F600-\U0001F64F"
        u"\U0001F300-\U0001F5FF"
        u"\U0001F680-\U0001F6FF"
        u"\U0001F1E0-\U0001F1FF"
        u"\U00002500-\U00002BEF"
        u"\U00002702-\U000027B0"
        u"\U000024C2-\U0001F251"
        u"\u2640-\u2642"
        u"\u2600-\u2B55"
        u"\u200d"
        u"\u23cf"
        u"\u23e9"
        u"\u231a"
        u"\ufe0f"
        u"\u3030"
        "]+", flags=re.UNICODE
    )
    return emoji_pattern.sub("", text).strip()


def auto_col_width(series: pd.Series, header: str) -> float:
    max_len = max(
        series.astype(str).map(len).max(),
        len(str(header))
    )
    width = max_len * 1.18 + 2
    return max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, width))


def is_numeric_col(series: pd.Series) -> bool:
    try:
        pd.to_numeric(series.dropna())
        return True
    except (ValueError, TypeError):
        return False


def convert(csv_path: str, output_path: str = None):
    csv_path = Path(csv_path)
    if not csv_path.exists():
        print(f"Error: File not found — {csv_path}")
        sys.exit(1)

    output_path = Path(output_path) if output_path else csv_path.with_suffix(".xlsx")
    stem = strip_emojis(csv_path.stem)

    # ── Load CSV ──────────────────────────────────────────────────────────────
    try:
        df = pd.read_csv(csv_path, dtype=str)
    except Exception:
        # Fallback: skip malformed rows
        df = pd.read_csv(csv_path, dtype=str, on_bad_lines="skip", engine="python")
    df.columns = [strip_emojis(str(c)) for c in df.columns]
    df = df.fillna("")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = stem[:31]  # Excel sheet name limit

    num_cols = len(df.columns)
    last_col = get_column_letter(num_cols)

    # ── Title banner (row 1) ──────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = stem
    title_cell.font = Font(name=FONT_NAME, bold=True, size=13, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=DARK_NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = TITLE_HEIGHT

    # ── Column headers (row 2) ────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(name=FONT_NAME, bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = HEADER_HEIGHT

    # ── Data rows (from row 3) ────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        bg = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
        for col_idx, value in enumerate(row, start=1):
            col_name = df.columns[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=strip_emojis(str(value)))
            cell.font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = THIN_BORDER
            numeric = is_numeric_col(df[col_name])
            cell.alignment = Alignment(
                horizontal="right" if numeric else "left",
                vertical="center",
                wrap_text=False
            )
        ws.row_dimensions[row_idx].height = ROW_HEIGHT

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        width = auto_col_width(df[col_name], col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze & filter ───────────────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}2"

    # ── Print setup ───────────────────────────────────────────────────────────
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:2"

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python convert_csv_to_xlsx.py input.csv [output.xlsx]")
        sys.exit(1)
    csv_in  = sys.argv[1]
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else None
    convert(csv_in, xlsx_out)
